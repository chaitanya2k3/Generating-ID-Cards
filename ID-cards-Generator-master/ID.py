from PIL import Image, ImageDraw, ImageFont
from tkinter.filedialog import askopenfile
import pyqrcode
from openpyxl import load_workbook
import subprocess
import sys
# os package to run operating system commands through python(pre installed)
import os


def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])


#install('pyqrcode')
#install('openpyxl')    
#install('Pillow')
install('pypng')


# openpyxl package for accessing excel sheets

# pyqrcode package to generate qr codes


try:
    file = askopenfile(title='Select the Workbook', mode='r', filetypes=[
        ('Microsoft Excel', '.xlsx .xlsx .xlsm .xltx .xltm')])
except:
    {}

if file is not None:
    dirpath = os.path.dirname(file.name)
    wbname = os.path.basename(file.name)
else:
    sys.exit()

# filepath variable contains the path of d members info excel sheet
# r'' makes d string in raw format so dat special characters like \n,\t will not b taken into consideration
filepath = file.name

# dirpath variable contains the address of location where we r going to save th qr codes n txt file
# folder is named as QR_Codes
dirpath = dirpath+'/'+'QR_Codes'

allIdsPath = dirpath+'/All_IDs'

# os.mkdir(filepath) is used to create a folder/dir where we r saving d qr codes
# its written in try-except block so that if the folder is already present den exception is generated
# bt will nt cause any problem to d execution of d program
try:
    os.mkdir(dirpath)
    os.mkdir(allIdsPath)
except:
    {}

# wb variable stores the workbook of d excel file opened
# load_workbook opens d workbook
wb = load_workbook(filepath)

print(end='\n\n')
print(wbname, end='\n\n')

# for every worksheet(ws) in d opened workbook(wb):
for ws in wb:

    print(ws.title)

    # wspath stores d path of another inside folder wid d name of the worksheet
    # .strip() is used to remove extra spaces at d ends of a string
    wspath = dirpath+'/'+ws.title.strip()

    #  it creates another folder/dir inside QR_Codes folder wid d name of d worksheet ws
    try:
        os.mkdir(wspath)
    except:
        {}

    # for every row in d worksheet which starts from 3 to 60
    # as d info of members starts from row no 3
    # at max no. of students could be 60
    for r in range(1, 190):
        # cell variable selects the cell which is placed at location row=r n column=1
        cell = ws.cell(row=r, column=1)

        # if the first cell(sl_no cell) is empty the loop will break
        # cell.value gives d data inside d cell
        if cell.value is None:
            break

        # rno,name,pno,mail are variables which store d respective data of a member of row=r
        # .strip() is used to remove extra spaces at d ends of a string
        l=[]
        rno = str(ws.cell(row=r, column=1).value).strip()
        #name = str(ws.cell(row=r, column=3).value).strip().replace('.',' ')
        fname = str(ws.cell(row=r, column=3).value).strip().replace('.',' ')
        lname = str(ws.cell(row=r, column=4).value).strip().replace('.',' ')
        #pno = str(int(ws.cell(row=r, column=6).value)).strip()
        pno = str(ws.cell(row=r, column=7).value).strip()
        mail = str(ws.cell(row=r, column=2).value).strip()
        l.extend([rno,fname,lname,pno,mail])
        if 'None' in l:
            continue

        # acm_id stores d acm id of a member
        # acm_id = yr(2 digits) - ACM – 0/1(1 digit) – Branch(2 alpha) – Roll no. (last 2 digits)
        acm_id = ''
        print(rno)
        print(len(rno))
        if(rno[4] == '5'):
            acm_id = str(int(rno[:2])-1)+'ACM1'
            #  21ACM1

        else:
            acm_id = rno[:2]+'ACM0'
            # 21ACM0

        acm_id += 'IT'+rno[8:]
        # 21ACM(0/1)IT22

        # eachmembertetepath variable stores the path address where we r going to store d qrcode
        eachmemberpath = wspath+'/'+acm_id

        # it creates a folder with name as acm_id inside d sheet name folder
        try:
            os.mkdir(eachmemberpath)
        except:
            {}

        # qr_info variable stores information to be stored in d qr code
        # roll,name,pno,mail are concatenated with ','
        qr_info = str(rno)+','+str(fname)+',' + \
            str(lname)+','+str(pno)+','+str(mail)

        # qr_code variable stores d qr code image
        # pyqrcode.create(info) function creates qr code wid d given info
        qr_code = pyqrcode.create(qr_info)

        qrpath = eachmemberpath+'/'+acm_id+'_qr.png'
        # idpath stores the path where the id is to be stored with the name as acm_id and extension .png
        idpath = eachmemberpath+'/'+acm_id+'_id.png'

        # this line saves d qr_code image in png format at imagepath location
        qr_code.png(qrpath, scale=7)

        qr_code = Image.open(qrpath)
        qr_code = qr_code.resize((945, 945))
        qr_code = qr_code.crop((72, 72, 873, 873))
        qr_code.save(qrpath)

        template = Image.open('sample.jpg')
        qr_code = Image.open(qrpath)
        template.paste(qr_code, (2850, 1440))
        # template.show()

        draw = ImageDraw.Draw(template)
        id_font = ImageFont.truetype('blogger-sans.medium.ttf', 150)
        name_font = ImageFont.truetype('alegreya-sans-sc.bold.ttf', 165)
        pno_font = ImageFont.truetype('AVGARDD_2.TTF', 130)

        draw.text((255, 1005), 'ACM ID : ' + acm_id, fill=(0, 0, 0, 255), font=id_font)
        draw.text((255, 1536), fname.upper()+' '+lname.upper(), fill=(0, 0, 0, 255), font=name_font)
        draw.text((324, 1755), '+91 '+pno, fill=(0, 0, 0, 255), font=pno_font)
        # template.show()
        template.save(idpath)
        template.save(allIdsPath+'/'+acm_id+'.png')

        # fp variable stores d address of a txt file containg d qr_info which is to be saved beside d qr code image
        fp = eachmemberpath+'/'+acm_id+'.txt'

        # file_acm stores a file opened at fp location in write mode
        # if d file already exists it overwrites it
        file_acm = open(fp, 'w')

        # lines variable contains d info to b stored in d file_acm file
        lines = [rno, '\n', fname, '\n', lname, '\n', pno, '\n', mail]

        # file.writelines(list of lines) function writes d given lines on d file_acm
        file_acm.writelines(lines)

        # file_acm is closed
        file_acm.close()

        # it continues with d next row

    # it continues with d next sheet in wb
        print(fname, lname, 'Done.')
    # print(ws.title, 'Sheet', 'Done.')
    print()
# after d job is done ... a Done statement is printed in the logs...!!!
print(wbname, 'WorkBook Done.', end='\n\n')